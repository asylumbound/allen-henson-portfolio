#!/usr/bin/env python3
"""
Allen Henson Portfolio - Product Catalog Spreadsheet Generator
Creates a comprehensive Excel spreadsheet with all products, variants, pricing, and assets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import os

# Theme Configuration - Elegant Black (matches cinematic noir aesthetic)
THEME = {
    'primary': '2D2D2D',
    'light': 'E5E5E5',
    'accent': 'C9A227',  # Gold accent to match site
    'chart_colors': ['2D2D2D', '4A4A4A', '6B6B6B', '8C8C8C', 'ADADAD', 'CFCFCF'],
}

SERIF_FONT = 'Georgia'
SANS_FONT = 'Calibri'

# Border styles
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Complete product catalog with variants
PRODUCTS = [
    # Page 1 Products (1-24)
    {
        "id": 1,
        "slug": "abscond-box-set",
        "name": "[IN PRODUCTION] LIMITED RUN - ABSCOND BOX SET Vol I-VI",
        "category": "boxset",
        "description": "Six-volume journey through France, Morocco, Italy, Greece, Prague & The Escape, and Los Angeles & The Virus",
        "status": "in_production",
        "image": "/images/sales/abscond-box-set.jpg",
        "variants": [
            {"name": "Complete Box Set", "price": 60000, "sku": "ABS-BOX-001"}
        ]
    },
    {
        "id": 2,
        "slug": "abscond-series",
        "name": "[PRESALE] ABSCOND - THE SERIES",
        "category": "book",
        "description": "The complete six-volume series available for presale",
        "status": "presale",
        "image": "/images/sales/abscond-series.png",
        "variants": [
            {"name": "Complete Series", "price": 51000, "sku": "ABS-SER-001"}
        ]
    },
    {
        "id": 3,
        "slug": "abscond-vol1-france",
        "name": "[PRESALE] ABSCOND - VOL I - FRANCE (I of VI)",
        "category": "book",
        "description": "Book One of Six in the Abscond Serial",
        "status": "presale",
        "image": "/images/sales/abscond-vol1-france.png",
        "variants": [
            {"name": "Standard Edition", "price": 5000, "sku": "ABS-V1-001"}
        ]
    },
    {
        "id": 4,
        "slug": "editorial-on-the-run",
        "name": "Editorial on the Run",
        "category": "book",
        "description": "Hell-on-wheels tour around U.S. wildernesses. 20,000 miles in 5 months.",
        "status": "available",
        "image": "/images/sales/editorial-on-the-run.png",
        "variants": [
            {"name": "Hardback Edition", "price": 5000, "sku": "EOTR-001"}
        ]
    },
    {
        "id": 5,
        "slug": "editorial-on-the-rocks",
        "name": "Editorial on the Rocks",
        "category": "book",
        "description": "The Photography of Allen Henson",
        "status": "available",
        "image": "/images/sales/editorial-on-the-rocks.png",
        "variants": [
            {"name": "Hardback Edition", "price": 5000, "sku": "EOTR-002"}
        ]
    },
    {
        "id": 6,
        "slug": "tour-de-eiffel",
        "name": "Tour de Eiffel + Mannequin Ingrat [MI001-045]",
        "category": "print",
        "description": "Limited edition print from the Mannequin Ingrat series",
        "status": "available",
        "image": "/images/sales/tour-de-eiffel.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 269000, "sku": "TDE-11X17"},
            {"name": "24\"X36\"", "price": 555000, "sku": "TDE-24X36"}
        ]
    },
    {
        "id": 7,
        "slug": "il-pantheon",
        "name": "Il Pantheon a Mezzanotte - [PAN001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/il-pantheon.jpg",
        "variants": [
            {"name": "Standard", "price": 500000, "sku": "PAN-001"}
        ]
    },
    {
        "id": 8,
        "slug": "sarah-in-london",
        "name": "Sarah in London [SL-001-050]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/sarah-in-london.jpg",
        "variants": [
            {"name": "Standard", "price": 110000, "sku": "SL-001"}
        ]
    },
    {
        "id": 9,
        "slug": "sword-bordeaux-v2",
        "name": "The Sword of Bordeaux v2of3 [2 SoB001-035]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/sword-bordeaux-v2.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 375000, "sku": "SOB2-11X17"},
            {"name": "24\"X36\"", "price": 475000, "sku": "SOB2-24X36"}
        ]
    },
    {
        "id": 10,
        "slug": "raffaella-tresor",
        "name": "Raffaella Trésor - Un Incrocio a Milano [RTit001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/raffaella-tresor.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 150000, "sku": "RT-11X17"},
            {"name": "24\"X36\"", "price": 240000, "sku": "RT-24X36"}
        ]
    },
    {
        "id": 11,
        "slug": "sacrilege-toulouse",
        "name": "Sacrilège à Toulouse - Mannequin in Toulouse II [SATII001-055]",
        "category": "print",
        "description": "Limited edition print from the Mannequin series",
        "status": "available",
        "image": "/images/sales/sacrilege-toulouse.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 555000, "sku": "SAT2-11X17"},
            {"name": "24\"X36\"", "price": 770000, "sku": "SAT2-24X36"}
        ]
    },
    {
        "id": 12,
        "slug": "mi-trevi",
        "name": "Mi Trevi! - Mannequin in Roma II [MTII001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/mi-trevi.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 190000, "sku": "MT2-11X17"},
            {"name": "24\"X36\"", "price": 350000, "sku": "MT2-24X36"}
        ]
    },
    {
        "id": 13,
        "slug": "sword-bordeaux-v1",
        "name": "The Sword of Bordeaux v1of3 [SoB001-035]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/sword-bordeaux-v1.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 450000, "sku": "SOB1-11X17"},
            {"name": "24\"X36\"", "price": 650000, "sku": "SOB1-24X36"}
        ]
    },
    {
        "id": 14,
        "slug": "sarina-thai",
        "name": "Sarina Thai in Grand Central 2015 [STG001-045]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/sarina-thai.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 595000, "sku": "STG-11X17"},
            {"name": "24\"X36\"", "price": 900000, "sku": "STG-24X36"}
        ]
    },
    {
        "id": 15,
        "slug": "entourage-pantheon-vii",
        "name": "Entourage al Pantheon VII [EAPII001-150] + Verisart Cert",
        "category": "print",
        "description": "Limited edition print with Verisart Certificate",
        "status": "available",
        "image": "/images/sales/entourage-pantheon-vii.png",
        "variants": [
            {"name": "Standard + Cert", "price": 500000, "sku": "EAP7-001"}
        ]
    },
    {
        "id": 16,
        "slug": "entourage-pantheon",
        "name": "Entourage al Pantheon [EAP001-150] + Verisart Cert",
        "category": "print",
        "description": "Limited edition print with Verisart Certificate",
        "status": "available",
        "image": "/images/sales/entourage-pantheon.png",
        "variants": [
            {"name": "Standard + Cert", "price": 500000, "sku": "EAP-001"}
        ]
    },
    {
        "id": 17,
        "slug": "agency-fees",
        "name": "AGENCY FEE'S 07JUNE2021",
        "category": "print",
        "description": "Limited edition print",
        "status": "sold_out",
        "image": "/images/sales/agency-fees.jpg",
        "variants": [
            {"name": "Standard", "price": 860000, "sku": "AF-001"}
        ]
    },
    {
        "id": 18,
        "slug": "tour-eiffel-paris",
        "name": "Tour Eiffel - Paris [TEII001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/tour-eiffel-paris.jpg",
        "variants": [
            {"name": "11\"X17\"", "price": 245000, "sku": "TE2-11X17"},
            {"name": "24\"X36\"", "price": 275000, "sku": "TE2-24X36"}
        ]
    },
    {
        "id": 19,
        "slug": "sunbathers-miami",
        "name": "Sunbathers in Miami Beach - 2014",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/sunbathers-miami.jpg",
        "variants": [
            {"name": "Standard", "price": 270000, "sku": "SMB-001"}
        ]
    },
    {
        "id": 20,
        "slug": "editorial-silver-gelatin",
        "name": "Editorial on the Run (Silver Gelatin FRAMED) - [OTR001-015L]",
        "category": "print",
        "description": "Silver Gelatin fiber print, framed",
        "status": "available",
        "image": "/images/sales/editorial-silver-gelatin.jpg",
        "variants": [
            {"name": "Framed", "price": 520000, "sku": "EOTR-SG-001"}
        ]
    },
    {
        "id": 21,
        "slug": "journal-44",
        "name": "Journal # 44 [The EXILE Journal] - Allen Henson",
        "category": "book",
        "description": "Personal journal documentation",
        "status": "sold_out",
        "image": "/images/sales/journal-44.jpg",
        "variants": [
            {"name": "Original", "price": 1500000, "sku": "J44-001"}
        ]
    },
    {
        "id": 22,
        "slug": "zines",
        "name": "The Zines, LASCIVIOUS + PARAPHILIA",
        "category": "book",
        "description": "Two-zine set",
        "status": "sold_out",
        "image": "/images/sales/zines.jpg",
        "variants": [
            {"name": "Two-Zine Set", "price": 9500, "sku": "ZINE-001"}
        ]
    },
    {
        "id": 23,
        "slug": "leaving-mondrian",
        "name": "Leaving the Mondrian - Miami Beach",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/leaving-mondrian.jpg",
        "variants": [
            {"name": "Standard", "price": 1399999, "sku": "LM-001"}
        ]
    },
    {
        "id": 24,
        "slug": "girl-smoking-coral",
        "name": "Girl smoking on Coral II - Miami [GSC001-020]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/girl-smoking-coral.jpg",
        "variants": [
            {"name": "Standard", "price": 970000, "sku": "GSC-001"}
        ]
    },
    # Page 2 Products (25-48)
    {
        "id": 25,
        "slug": "odlh-set",
        "name": "ODLH SET",
        "category": "print",
        "description": "Limited edition set",
        "status": "sold_out",
        "image": "/images/sales/odlh_set",
        "variants": [
            {"name": "Contact for Price", "price": 0, "sku": "ODLH-001"}
        ]
    },
    {
        "id": 26,
        "slug": "anna-oakley-silver-gelatin",
        "name": "Anna Oakley (Silver Gelatin) - [AO001-015L]",
        "category": "print",
        "description": "Silver Gelatin fiber print",
        "status": "available",
        "image": "/images/sales/anna-oakley-silver-gelatin",
        "variants": [
            {"name": "Standard", "price": 1290000, "sku": "AO-SG-001"}
        ]
    },
    {
        "id": 27,
        "slug": "corset-en-metal",
        "name": "Corset en Métal [CEMII001-015]",
        "category": "print",
        "description": "Limited edition print from the Corset series",
        "status": "available",
        "image": "/images/sales/Corset_en_Metal",
        "variants": [
            {"name": "Standard", "price": 1510000, "sku": "CEM-001"}
        ]
    },
    {
        "id": 28,
        "slug": "rudy-reyes-24x36",
        "name": "Rudy Reyes 24\"X36\"",
        "category": "print",
        "description": "Portrait of Rudy Reyes",
        "status": "available",
        "image": "/images/sales/rudy-reyes-24x36",
        "variants": [
            {"name": "11\"X17\"", "price": 655000, "sku": "RR-11X17"},
            {"name": "24\"X36\"", "price": 710000, "sku": "RR-24X36"}
        ]
    },
    {
        "id": 29,
        "slug": "rudy-reyes-ii",
        "name": "Rudy Reyes II",
        "category": "print",
        "description": "Portrait of Rudy Reyes",
        "status": "available",
        "image": "/images/sales/Rudy_Reyes_II",
        "variants": [
            {"name": "11\"X17\"", "price": 120000, "sku": "RR2-11X17"},
            {"name": "24\"X36\"", "price": 530000, "sku": "RR2-24X36"}
        ]
    },
    {
        "id": 30,
        "slug": "girl-coal-ny",
        "name": "Girl + Coal NY 2015 [GC001-015]",
        "category": "print",
        "description": "New York City, 2015",
        "status": "available",
        "image": "/images/sales/Girl_Coal_NY_2015_GC001_015",
        "variants": [
            {"name": "11\"X17\"", "price": 530000, "sku": "GC-11X17"},
            {"name": "24\"X36\"", "price": 770000, "sku": "GC-24X36"}
        ]
    },
    {
        "id": 31,
        "slug": "foro-romano",
        "name": "Foro Romano - Rome Italy [LP07]",
        "category": "print",
        "description": "Rome, Italy",
        "status": "available",
        "image": "/images/sales/Foro_Romano_Rome_Italy_LP07",
        "variants": [
            {"name": "11\"X17\"", "price": 310000, "sku": "FR-11X17"},
            {"name": "24\"X36\"", "price": 495000, "sku": "FR-24X36"}
        ]
    },
    {
        "id": 32,
        "slug": "journal-22",
        "name": "Journal # 22 - Allen Henson",
        "category": "book",
        "description": "Personal journal documentation",
        "status": "sold_out",
        "image": "/images/sales/journal-22-allen-henson",
        "variants": [
            {"name": "Original", "price": 800000, "sku": "J22-001"}
        ]
    },
    {
        "id": 33,
        "slug": "anna-lisa-sequoia",
        "name": "Anna Lisa in Sequoiadendron Giganteum",
        "category": "print",
        "description": "Shot among the giant sequoias",
        "status": "available",
        "image": "/images/sales/anna_lisa_in_sequoiadendron_giganteum",
        "variants": [
            {"name": "11\"X17\"", "price": 410000, "sku": "ALS-11X17"},
            {"name": "24\"X36\"", "price": 670000, "sku": "ALS-24X36"}
        ]
    },
    {
        "id": 34,
        "slug": "what-we-left-paris",
        "name": "What we left in Paris [LIP001-015]",
        "category": "print",
        "description": "Paris, France",
        "status": "available",
        "image": "/images/sales/what-we-left-in-paris",
        "variants": [
            {"name": "11\"X17\"", "price": 1550000, "sku": "LIP-11X17"},
            {"name": "24\"X36\"", "price": 1700000, "sku": "LIP-24X36"}
        ]
    },
    {
        "id": 35,
        "slug": "ipseity",
        "name": "Ipseity - [IPS001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/ipseity-ips001-015",
        "variants": [
            {"name": "Standard", "price": 540000, "sku": "IPS-001"}
        ]
    },
    {
        "id": 36,
        "slug": "mouvement-paris",
        "name": "Mouvement Paris [MV001-015]",
        "category": "print",
        "description": "Paris, France - capturing movement",
        "status": "available",
        "image": "/images/sales/Mouvement_Paris_MV001-015",
        "variants": [
            {"name": "11\"X17\"", "price": 610000, "sku": "MV-11X17"},
            {"name": "24\"X36\"", "price": 940000, "sku": "MV-24X36"}
        ]
    },
    {
        "id": 37,
        "slug": "burlesque-ny-2015",
        "name": "Burlesque - New York 2015",
        "category": "print",
        "description": "New York City, 2015",
        "status": "available",
        "image": "/images/sales/burlesque-new-york-2015",
        "variants": [
            {"name": "Standard", "price": 780000, "sku": "BUR-001"}
        ]
    },
    {
        "id": 38,
        "slug": "walk-to-cafe-paris",
        "name": "a walk to the Cafe - Paris June [CAFE001-015]",
        "category": "print",
        "description": "Paris, France",
        "status": "available",
        "image": "/images/sales/a_walk_to_the_Cafe_-_Paris_June_[CAFE001-015]",
        "variants": [
            {"name": "11\"X17\"", "price": 995000, "sku": "CAFE-11X17"},
            {"name": "24\"X36\"", "price": 1355000, "sku": "CAFE-24X36"}
        ]
    },
    {
        "id": 39,
        "slug": "helene-traasavik-i",
        "name": "Helene Traasavik I - Los Angeles [HTI001-015]",
        "category": "print",
        "description": "Los Angeles",
        "status": "available",
        "image": "/images/sales/helene-traasavik-i-los-angeles",
        "variants": [
            {"name": "Standard", "price": 350000, "sku": "HT1-001"}
        ]
    },
    {
        "id": 40,
        "slug": "gun-rights-la",
        "name": "¿Gun Rights? - Los Angeles [GRL001-015]",
        "category": "print",
        "description": "Los Angeles",
        "status": "available",
        "image": "/images/sales/Gun_Rights_Los_Angeles_2012",
        "variants": [
            {"name": "Standard", "price": 1299000, "sku": "GR-001"}
        ]
    },
    {
        "id": 41,
        "slug": "odeon-herodes-atticus",
        "name": "The Odeon of Herodes Atticus - Mannequin [HA001-015]",
        "category": "print",
        "description": "Athens, Greece",
        "status": "available",
        "image": "/images/sales/The_Odeon_of_Herodes_Atticus_Mannequin",
        "variants": [
            {"name": "11\"X17\"", "price": 1515000, "sku": "OHA-11X17"},
            {"name": "24\"X36\"", "price": 1650000, "sku": "OHA-24X36"}
        ]
    },
    {
        "id": 42,
        "slug": "mi-trevi-skye-roma",
        "name": "Mi Trevi! - Skye in Roma [MT001-015]",
        "category": "print",
        "description": "Rome, Italy - Trevi Fountain",
        "status": "available",
        "image": "/images/sales/mi-trevi-skye-in-roma-mt001-015",
        "variants": [
            {"name": "11\"X17\"", "price": 1175000, "sku": "MTS-11X17"},
            {"name": "24\"X36\"", "price": 1525000, "sku": "MTS-24X36"}
        ]
    },
    {
        "id": 43,
        "slug": "girl-on-coral",
        "name": "Girl on Coral [GCM001-015]",
        "category": "print",
        "description": "Miami, Florida",
        "status": "available",
        "image": "/images/sales/girl-on-coral",
        "variants": [
            {"name": "11\"X17\"", "price": 1770000, "sku": "GOC-11X17"},
            {"name": "24\"X36\"", "price": 2245000, "sku": "GOC-24X36"}
        ]
    },
    {
        "id": 44,
        "slug": "ryan-hunter-miami",
        "name": "Ryan Hunter - Miami - Venetian",
        "category": "print",
        "description": "Miami Beach, Florida",
        "status": "available",
        "image": "/images/sales/ryan-hunter-miami-venetian-2014",
        "variants": [
            {"name": "Standard", "price": 1170000, "sku": "RH-001"}
        ]
    },
    {
        "id": 45,
        "slug": "sarina-flatiron",
        "name": "Sarina Flatiron Building - NYC [SFB001-015]",
        "category": "print",
        "description": "New York City - Flatiron Building",
        "status": "available",
        "image": "/images/sales/Sarian_Flatiron_Building_NYC_2015",
        "variants": [
            {"name": "Standard", "price": 1290000, "sku": "SFB-001"}
        ]
    },
    {
        "id": 46,
        "slug": "mannequin-mast-barcelona",
        "name": "Mannequin on the Mast en Barcelona [SB10]",
        "category": "print",
        "description": "Barcelona, Spain",
        "status": "available",
        "image": "/images/sales/Mannequin_on_the_Mast_en_Barcelona_SB10",
        "variants": [
            {"name": "11\"X17\"", "price": 700000, "sku": "MMB-11X17"},
            {"name": "24\"X36\"", "price": 820000, "sku": "MMB-24X36"}
        ]
    },
    {
        "id": 47,
        "slug": "colosseum-rome",
        "name": "Colosseum - Rome [C99]",
        "category": "print",
        "description": "Rome, Italy",
        "status": "available",
        "image": "/images/sales/Colosseum_Rome_C99",
        "variants": [
            {"name": "Standard", "price": 235000, "sku": "COL-001"}
        ]
    },
    {
        "id": 48,
        "slug": "room-102-access",
        "name": "ROOM 102 ACCESS (DISCONTINUED)",
        "category": "access",
        "description": "Discontinued access pass",
        "status": "available",
        "image": "/images/sales/room_102_access",
        "variants": [
            {"name": "Digital Access", "price": 25000, "sku": "R102-001"}
        ]
    },
    # Page 3 Products (49-72)
    {
        "id": 49,
        "slug": "sacrilege-toulouse-skye",
        "name": "Sacrilège à Toulouse - Skye in Toulouse [SAT001-015]",
        "category": "print",
        "description": "Toulouse, France",
        "status": "available",
        "image": "/images/sales/Sacrilege_a_Toulouse_Skye_in_Toulouse_SAT001_015",
        "variants": [
            {"name": "11\"X17\"", "price": 2250000, "sku": "SATS-11X17"},
            {"name": "24\"X36\"", "price": 2350000, "sku": "SATS-24X36"}
        ]
    },
    {
        "id": 50,
        "slug": "emily-shephard-bisjoux",
        "name": "Emily Shephard in BISJOUX II [ESBi001-015]",
        "category": "print",
        "description": "BISJOUX series",
        "status": "available",
        "image": "/images/sales/Emily-Shephard-in-BISJOUX-II-ESBi001-015",
        "variants": [
            {"name": "Standard", "price": 1140000, "sku": "ESB-001"}
        ]
    },
    {
        "id": 51,
        "slug": "good-morning-paris",
        "name": "Tour de Eiffel + Mannequin Ingrat [2MI001-045] (Good Morning Paris!)",
        "category": "print",
        "description": "Paris, France - Mannequin Ingrat series",
        "status": "available",
        "image": "/images/sales/skye-eiffel-tower-good-morning-paris",
        "variants": [
            {"name": "Standard", "price": 2250000, "sku": "GMP-001"}
        ]
    },
    {
        "id": 52,
        "slug": "editorial-bundle",
        "name": "Editorial on the Run + Editorial on the Rocks",
        "category": "book",
        "description": "Bundle of both Editorial books",
        "status": "available",
        "image": "/images/sales/editorial_on_the_run_and_editorial_on_the_rocks",
        "variants": [
            {"name": "Two-Book Bundle", "price": 9000, "sku": "EOTR-BUN"}
        ]
    },
    {
        "id": 53,
        "slug": "journal-23",
        "name": "Journal # 23 - Allen Henson",
        "category": "book",
        "description": "Personal journal documentation",
        "status": "sold_out",
        "image": "/images/sales/Journal_23_Allen_Henson",
        "variants": [
            {"name": "Original", "price": 900000, "sku": "J23-001"}
        ]
    },
    {
        "id": 54,
        "slug": "karyna-studio-ny",
        "name": "Karyna - Studio N.Y. [KAS001-015]",
        "category": "print",
        "description": "New York City",
        "status": "available",
        "image": "/images/sales/Karyna-Studio-NY-KAS001-015",
        "variants": [
            {"name": "Standard", "price": 630000, "sku": "KSN-001"}
        ]
    },
    {
        "id": 55,
        "slug": "bespoke-camera-handles",
        "name": "Bespoke Wooden Camera Handles by Allen Henson",
        "category": "accessory",
        "description": "Handcrafted wooden camera handles",
        "status": "sold_out",
        "image": "/images/sales/bespoke-wooden-camera-handles-by-allen-henson",
        "variants": [
            {"name": "Handmade", "price": 35000, "sku": "BCH-001"}
        ]
    },
    {
        "id": 56,
        "slug": "karyna-on-dock",
        "name": "Karyna on Dock - [KKD001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/Karyna_on_Dock",
        "variants": [
            {"name": "Standard", "price": 1270000, "sku": "KOD-001"}
        ]
    },
    {
        "id": 57,
        "slug": "kara-gibson-la-ii",
        "name": "Kara Gibson - A.H. Studio L.A. II 2012",
        "category": "print",
        "description": "Los Angeles, 2012",
        "status": "available",
        "image": "/images/sales/Kara_Gibson_AH_Studio_LA_II_2012",
        "variants": [
            {"name": "Standard", "price": 1300000, "sku": "KG-001"}
        ]
    },
    {
        "id": 58,
        "slug": "leia-contois-la",
        "name": "Leia Contois - Los Angeles [LCL001-015]",
        "category": "print",
        "description": "Los Angeles",
        "status": "available",
        "image": "/images/sales/leia-contois-los-angeles-2012",
        "variants": [
            {"name": "Standard", "price": 1250000, "sku": "LC-001"}
        ]
    },
    {
        "id": 59,
        "slug": "pantheon-roma-2015",
        "name": "Pantheon - Roma 2015",
        "category": "print",
        "description": "Rome, Italy",
        "status": "sold_out",
        "image": "/images/sales/Pantheon-Roma-2015",
        "variants": [
            {"name": "Standard", "price": 140000, "sku": "PR-001"}
        ]
    },
    {
        "id": 60,
        "slug": "arc-de-triomphe",
        "name": "Arc de Triomphe - Paris",
        "category": "print",
        "description": "Paris, France",
        "status": "available",
        "image": "/images/sales/arc-de-triomphe-paris",
        "variants": [
            {"name": "Standard", "price": 170000, "sku": "ADT-001"}
        ]
    },
    {
        "id": 61,
        "slug": "portrait-girl-ny",
        "name": "A Portrait of a Girl - NY [PX001-015]",
        "category": "print",
        "description": "New York",
        "status": "available",
        "image": "/images/sales/a-portrait-of-a-girl-ny-px001-015",
        "variants": [
            {"name": "Standard", "price": 500000, "sku": "PG-001"}
        ]
    },
    {
        "id": 62,
        "slug": "london-big-ben",
        "name": "London - Big Ben [L001-015]",
        "category": "print",
        "description": "London, UK",
        "status": "available",
        "image": "/images/sales/london-big-ben-l001-015",
        "variants": [
            {"name": "Standard", "price": 250000, "sku": "LBB-001"}
        ]
    },
    {
        "id": 63,
        "slug": "cate-underwood-manhattan",
        "name": "Cate Underwood - Manhattan 2014",
        "category": "print",
        "description": "Manhattan, New York",
        "status": "available",
        "image": "/images/sales/cate-underwood-manhattan-2014",
        "variants": [
            {"name": "Standard", "price": 390000, "sku": "CU-001"}
        ]
    },
    {
        "id": 64,
        "slug": "tika-camaj-miami",
        "name": "Tika Camaj - Miami - Venetian 2014",
        "category": "print",
        "description": "Miami Beach, Florida",
        "status": "sold_out",
        "image": "/images/sales/Tika_Camaj_Miami_Venetian_2014",
        "variants": [
            {"name": "Standard", "price": 150000, "sku": "TC-001"}
        ]
    },
    {
        "id": 65,
        "slug": "data-licensing",
        "name": "Data Licensing",
        "category": "license",
        "description": "License data for commercial use",
        "status": "available",
        "image": "/images/sales/Data_Licensing",
        "variants": [
            {"name": "Contact for Price", "price": 0, "sku": "DL-001"}
        ]
    },
    {
        "id": 66,
        "slug": "gianluca-di-sotto",
        "name": "Gianluca di Sotto - NYC / L.E.S",
        "category": "print",
        "description": "New York City, Lower East Side",
        "status": "available",
        "image": "/images/sales/gianluca_di_sotto_nyc_les_2014",
        "variants": [
            {"name": "Standard", "price": 350000, "sku": "GDS-001"}
        ]
    },
    {
        "id": 67,
        "slug": "karyna-brooklyn",
        "name": "Karyna - Brooklyn 2015",
        "category": "print",
        "description": "Brooklyn, New York",
        "status": "sold_out",
        "image": "/images/sales/Karyna-Brooklyn-2015",
        "variants": [
            {"name": "Standard", "price": 130000, "sku": "KB-001"}
        ]
    },
    {
        "id": 68,
        "slug": "sara-balint-fidi",
        "name": "Sara Balint - FiDi NYC",
        "category": "print",
        "description": "Financial District, New York City",
        "status": "available",
        "image": "/images/sales/sara-balint-fidi-nyc-2014",
        "variants": [
            {"name": "Standard", "price": 520000, "sku": "SB-001"}
        ]
    },
    {
        "id": 69,
        "slug": "another-4am-miami",
        "name": "Another 4 a.m. shoot - Miami Beach 2014",
        "category": "print",
        "description": "Miami Beach, Florida",
        "status": "sold_out",
        "image": "/images/sales/another-4-am-shoot-miami-beach-2014",
        "variants": [
            {"name": "Standard", "price": 70000, "sku": "A4AM-001"}
        ]
    },
    {
        "id": 70,
        "slug": "karyna-union-league-ii",
        "name": "Karyna - Union League Club II - 2015",
        "category": "print",
        "description": "New York City",
        "status": "available",
        "image": "/images/sales/Karyna-Union-League-Club-II-2015",
        "variants": [
            {"name": "Standard", "price": 220000, "sku": "KUL2-001"}
        ]
    },
    {
        "id": 71,
        "slug": "karyna-union-league",
        "name": "Karyna - Union League Club",
        "category": "print",
        "description": "New York City",
        "status": "available",
        "image": "/images/sales/karyna_union_league_club",
        "variants": [
            {"name": "Standard", "price": 270000, "sku": "KUL-001"}
        ]
    },
    {
        "id": 72,
        "slug": "karyna-grand-central",
        "name": "Karyna - Grand Central Station",
        "category": "print",
        "description": "Grand Central Station, New York City",
        "status": "available",
        "image": "/images/sales/Karyna-Grand-Central-Station",
        "variants": [
            {"name": "Standard", "price": 290000, "sku": "KGC-001"}
        ]
    },
    # Page 4 Products (73-81)
    {
        "id": 73,
        "slug": "kat-miami-beach",
        "name": "Kat - Miami Beach 2014",
        "category": "print",
        "description": "Miami Beach, Florida",
        "status": "available",
        "image": "/images/sales/kat-miami-beach-2014",
        "variants": [
            {"name": "Standard", "price": 150000, "sku": "KMB-001"}
        ]
    },
    {
        "id": 74,
        "slug": "shelby-carter-empire-state",
        "name": "Shelby Carter / Elizabeth Marxs - Empire State Building 2013 [1of5]",
        "category": "print",
        "description": "Empire State Building, New York City",
        "status": "available",
        "image": "/images/sales/Shelby-Carter-Elizabeth-Marxs-Empire-State-Building-2013-1of5",
        "variants": [
            {"name": "1 of 5", "price": 270000, "sku": "SCEM-001"}
        ]
    },
    {
        "id": 75,
        "slug": "helene-traasavik-ii",
        "name": "Helene Traasavik II - Los Angeles 2013",
        "category": "print",
        "description": "Los Angeles, 2013",
        "status": "available",
        "image": "/images/sales/helene-traasavik-ii-los-angeles-2013",
        "variants": [
            {"name": "Standard", "price": 170000, "sku": "HT2-001"}
        ]
    },
    {
        "id": 76,
        "slug": "burlesque-ii-ny",
        "name": "Burlesque II - New York 2015",
        "category": "print",
        "description": "New York City, 2015",
        "status": "available",
        "image": "/images/sales/burlesque-ii-new-york-2015",
        "variants": [
            {"name": "Standard", "price": 300000, "sku": "BUR2-001"}
        ]
    },
    {
        "id": 77,
        "slug": "laundry-day-la",
        "name": "Laundry Day - Los Angeles 2012",
        "category": "print",
        "description": "Los Angeles, 2012",
        "status": "sold_out",
        "image": "/images/sales/laundry-day-los-angeles-2012",
        "variants": [
            {"name": "Standard", "price": 95000, "sku": "LD-001"}
        ]
    },
    {
        "id": 78,
        "slug": "victorious-venetian",
        "name": "Victorious on Venetian Rooftop 2014 Miami Beach",
        "category": "print",
        "description": "Miami Beach, Florida",
        "status": "available",
        "image": "/images/sales/victorious-on-venetian-rooftop-2014-miami-beach",
        "variants": [
            {"name": "Standard", "price": 150500, "sku": "VV-001"}
        ]
    },
    {
        "id": 79,
        "slug": "karyna-soho-nyc",
        "name": "Karyna in SoHo NYC 2014",
        "category": "print",
        "description": "SoHo, New York City",
        "status": "sold_out",
        "image": "/images/sales/Karyna_in_SoHo_NYC_2014",
        "variants": [
            {"name": "Standard", "price": 65000, "sku": "KS-001"}
        ]
    },
    {
        "id": 80,
        "slug": "batch-a113",
        "name": "BATCH A113 pt1of2",
        "category": "print",
        "description": "Part 1 of 2",
        "status": "available",
        "image": "/images/sales/batch-a113-pt1of2",
        "variants": [
            {"name": "Standard", "price": 200000, "sku": "BA113-001"}
        ]
    },
    {
        "id": 81,
        "slug": "paon-au-greystone",
        "name": "Paon au Greystone [PAG001-015]",
        "category": "print",
        "description": "Limited edition print",
        "status": "available",
        "image": "/images/sales/paon-au-greystone-pag001-015",
        "variants": [
            {"name": "Standard", "price": 500000, "sku": "PAG-001"}
        ]
    },
]


def format_price(cents):
    """Format price from cents to dollars"""
    if cents == 0:
        return "Contact"
    return f"${cents / 100:,.0f}"


def create_spreadsheet():
    """Create the comprehensive product catalog spreadsheet"""
    wb = Workbook()
    
    # ===== OVERVIEW SHEET =====
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.sheet_view.showGridLines = False
    
    # Left margin
    ws_overview.column_dimensions['A'].width = 3
    
    # Title
    ws_overview['B2'] = "ALLEN HENSON PRODUCTIONS"
    ws_overview['B2'].font = Font(name=SERIF_FONT, size=22, bold=True, color=THEME['primary'])
    
    ws_overview['B3'] = "Product Catalog & Pricing Guide"
    ws_overview['B3'].font = Font(name=SERIF_FONT, size=14, color='666666')
    
    ws_overview['B4'] = f"Generated: February 2026 | 81 Products | Multiple Variants"
    ws_overview['B4'].font = Font(name=SANS_FONT, size=10, italic=True, color='999999')
    
    # Summary stats
    ws_overview['B7'] = "CATALOG SUMMARY"
    ws_overview['B7'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    total_products = len(PRODUCTS)
    available = sum(1 for p in PRODUCTS if p['status'] == 'available')
    sold_out = sum(1 for p in PRODUCTS if p['status'] == 'sold_out')
    presale = sum(1 for p in PRODUCTS if p['status'] == 'presale')
    in_production = sum(1 for p in PRODUCTS if p['status'] == 'in_production')
    products_with_variants = sum(1 for p in PRODUCTS if len(p['variants']) > 1)
    
    stats = [
        ("Total Products", total_products),
        ("Available", available),
        ("Sold Out", sold_out),
        ("Presale", presale),
        ("In Production", in_production),
        ("Products with Size Variants", products_with_variants),
    ]
    
    for i, (label, value) in enumerate(stats, start=9):
        ws_overview[f'B{i}'] = label
        ws_overview[f'B{i}'].font = Font(name=SANS_FONT, size=11)
        ws_overview[f'C{i}'] = value
        ws_overview[f'C{i}'].font = Font(name=SANS_FONT, size=11, bold=True)
    
    # Sheet index
    ws_overview['B17'] = "CONTENTS"
    ws_overview['B17'].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
    
    sheets = ["Overview", "Full Catalog", "Variants Detail", "By Category"]
    for i, sheet_name in enumerate(sheets, start=19):
        cell = ws_overview.cell(row=i, column=2, value=sheet_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(name=SANS_FONT, size=11, color=THEME['accent'], underline='single')
    
    ws_overview.column_dimensions['B'].width = 35
    ws_overview.column_dimensions['C'].width = 15
    
    # ===== FULL CATALOG SHEET =====
    ws_catalog = wb.create_sheet("Full Catalog")
    ws_catalog.sheet_view.showGridLines = False
    ws_catalog.column_dimensions['A'].width = 3
    
    # Title
    ws_catalog['B2'] = "FULL PRODUCT CATALOG"
    ws_catalog['B2'].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['primary'])
    
    # Headers
    headers = ["ID", "Slug", "Name", "Category", "Status", "Min Price", "Max Price", "Has Variants", "Image Path"]
    header_fill = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
    header_font = Font(name=SERIF_FONT, size=10, bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, start=2):
        cell = ws_catalog.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, product in enumerate(PRODUCTS, start=5):
        min_price = min(v['price'] for v in product['variants'])
        max_price = max(v['price'] for v in product['variants'])
        has_variants = "Yes" if len(product['variants']) > 1 else "No"
        
        data = [
            product['id'],
            product['slug'],
            product['name'],
            product['category'],
            product['status'],
            format_price(min_price),
            format_price(max_price) if max_price != min_price else "-",
            has_variants,
            product['image']
        ]
        
        for col, value in enumerate(data, start=2):
            cell = ws_catalog.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name=SANS_FONT, size=10)
            cell.border = thin_border
            
            # Status coloring
            if col == 6:  # Status column
                if value == 'sold_out':
                    cell.font = Font(name=SANS_FONT, size=10, color='C62828')
                elif value == 'presale':
                    cell.font = Font(name=SANS_FONT, size=10, color=THEME['accent'])
                elif value == 'in_production':
                    cell.font = Font(name=SANS_FONT, size=10, color='F57C00')
    
    # Column widths
    ws_catalog.column_dimensions['B'].width = 6   # ID
    ws_catalog.column_dimensions['C'].width = 30  # Slug
    ws_catalog.column_dimensions['D'].width = 55  # Name
    ws_catalog.column_dimensions['E'].width = 12  # Category
    ws_catalog.column_dimensions['F'].width = 12  # Status
    ws_catalog.column_dimensions['G'].width = 12  # Min Price
    ws_catalog.column_dimensions['H'].width = 12  # Max Price
    ws_catalog.column_dimensions['I'].width = 12  # Has Variants
    ws_catalog.column_dimensions['J'].width = 50  # Image Path
    
    # Freeze panes
    ws_catalog.freeze_panes = 'B5'
    
    # ===== VARIANTS DETAIL SHEET =====
    ws_variants = wb.create_sheet("Variants Detail")
    ws_variants.sheet_view.showGridLines = False
    ws_variants.column_dimensions['A'].width = 3
    
    # Title
    ws_variants['B2'] = "PRODUCT VARIANTS & PRICING"
    ws_variants['B2'].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['primary'])
    
    ws_variants['B3'] = "Detailed breakdown of all size options and pricing tiers"
    ws_variants['B3'].font = Font(name=SANS_FONT, size=11, italic=True, color='666666')
    
    # Headers
    variant_headers = ["Product ID", "Product Name", "Variant Name", "Price", "SKU"]
    for col, header in enumerate(variant_headers, start=2):
        cell = ws_variants.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows - one row per variant
    row_idx = 6
    for product in PRODUCTS:
        for variant in product['variants']:
            data = [
                product['id'],
                product['name'],
                variant['name'],
                format_price(variant['price']),
                variant['sku']
            ]
            
            for col, value in enumerate(data, start=2):
                cell = ws_variants.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name=SANS_FONT, size=10)
                cell.border = thin_border
            
            row_idx += 1
    
    # Column widths
    ws_variants.column_dimensions['B'].width = 12  # Product ID
    ws_variants.column_dimensions['C'].width = 55  # Product Name
    ws_variants.column_dimensions['D'].width = 15  # Variant Name
    ws_variants.column_dimensions['E'].width = 15  # Price
    ws_variants.column_dimensions['F'].width = 15  # SKU
    
    # Freeze panes
    ws_variants.freeze_panes = 'B6'
    
    # ===== BY CATEGORY SHEET =====
    ws_category = wb.create_sheet("By Category")
    ws_category.sheet_view.showGridLines = False
    ws_category.column_dimensions['A'].width = 3
    
    # Title
    ws_category['B2'] = "PRODUCTS BY CATEGORY"
    ws_category['B2'].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME['primary'])
    
    # Group products by category
    categories = {}
    for product in PRODUCTS:
        cat = product['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(product)
    
    row_idx = 5
    for category, products in sorted(categories.items()):
        # Category header
        ws_category.cell(row=row_idx, column=2, value=category.upper())
        ws_category.cell(row=row_idx, column=2).font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME['accent'])
        ws_category.cell(row=row_idx, column=2).fill = PatternFill(start_color=THEME['light'], end_color=THEME['light'], fill_type='solid')
        row_idx += 1
        
        # Table headers
        cat_headers = ["Name", "Status", "Price Range"]
        for col, header in enumerate(cat_headers, start=2):
            cell = ws_category.cell(row=row_idx, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row_idx += 1
        
        # Products in category
        for product in products:
            min_price = min(v['price'] for v in product['variants'])
            max_price = max(v['price'] for v in product['variants'])
            price_range = format_price(min_price)
            if max_price != min_price:
                price_range += f" - {format_price(max_price)}"
            
            data = [product['name'], product['status'], price_range]
            for col, value in enumerate(data, start=2):
                cell = ws_category.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name=SANS_FONT, size=10)
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 2  # Space between categories
    
    # Column widths
    ws_category.column_dimensions['B'].width = 55
    ws_category.column_dimensions['C'].width = 15
    ws_category.column_dimensions['D'].width = 25
    
    # Save workbook
    output_path = '/home/ubuntu/allen-henson-portfolio/docs/allen_henson_product_catalog.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Spreadsheet saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    create_spreadsheet()
